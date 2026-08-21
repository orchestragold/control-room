"""
pitch_targets sync engine.

Rebuilds the pitch_targets table from all three sources in order:
  1. HubSpot company cache  (hubspot_companies table)
  2. Festival Outreach XLSX (Dropbox — path in SPREADSHEET_DROPBOX_PATH config)
  3. pitch_queue.csv        (Dropbox App folder)

Spec: /Apps/PortalKnowledgeSync/Session G-1 Spec - pitch_targets.md

Public API
----------
  sync_pitch_targets(xlsx_bytes=None, csv_content=None) -> SyncResult

  Pass xlsx_bytes/csv_content explicitly in tests to avoid Dropbox calls.
  Pass None (default) to download live from Dropbox.
"""

from __future__ import annotations

import io
import re
import sys
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Optional

from app.extensions import db
from app.models.hubspot_cache import HubSpotCompany
from app.models.pitch_target import PitchTarget
from app.pitch_machine.stages import PMStage

# ── Spreadsheet constants ─────────────────────────────────────────────────────

SPREADSHEET_SHEET_NAME = 'Festival Outreach List'

# First matching alias (case-insensitive) wins per field.
_COL_ALIASES: dict[str, list[str]] = {
    'name':                ['Festival Name', 'Name', 'Target', 'Festival', 'Organization'],
    'status':              ['Status'],
    'hubspot_linked':      ['HubSpot Linked', 'Hubspot Linked', 'HubSpot ID', 'HS Linked'],
    'website':             ['Website', 'URL'],
    'pitch_type':          ['Pitch Type', 'Type'],
    'submission_deadline': ['Submission Deadline', 'Deadline', 'Sub Deadline'],
}

# "Yes (338507124464)" — extract the numeric HubSpot company ID
_HS_LINKED_RE = re.compile(r'Yes\s*\((\d+)\)', re.IGNORECASE)


# ── SyncResult ────────────────────────────────────────────────────────────────

@dataclass
class SyncResult:
    hubspot_count:       int  = 0
    spreadsheet_count:   int  = 0
    csv_count:           int  = 0
    conflict_count:      int  = 0
    not_a_fit_count:     int  = 0
    total:               int  = 0
    spreadsheet_skipped: bool = False
    warnings: list[str] = field(default_factory=list)

    def __str__(self) -> str:
        parts = [f'Total: {self.total}', f'HubSpot: {self.hubspot_count}']
        if self.spreadsheet_skipped:
            parts.append('Spreadsheet: skipped')
        else:
            parts.append(f'Spreadsheet: {self.spreadsheet_count}')
        parts += [
            f'CSV: {self.csv_count}',
            f'Conflicts: {self.conflict_count}',
            f'Not-a-fit: {self.not_a_fit_count}',
        ]
        return ' | '.join(parts)


# ── Name normalization ────────────────────────────────────────────────────────

def normalize_name(name: str) -> str:
    """Canonical key for name-based dedup: trimmed, lowercased."""
    return name.strip().lower()


# ── Stage mapping ─────────────────────────────────────────────────────────────

def spreadsheet_status_to_stage(raw: Optional[str]) -> str:
    """
    Map a raw spreadsheet Status value to a PMStage string.

    Unrecognized values → NEEDS_OUTREACH (never raise; the spreadsheet is
    continuously edited and will grow new free-text Status values over time).
    """
    if not raw:
        return PMStage.NEEDS_OUTREACH.value
    s = raw.strip()
    if s.startswith('Pitch Sent'):
        return PMStage.SENT.value
    if s.startswith('Pitch Scheduled'):
        return PMStage.QUEUED.value
    if s.startswith('No Outreach Needed'):
        return PMStage.DEPRIORITIZED.value
    if s.startswith('INACTIVE'):
        return PMStage.DECLINED.value
    return PMStage.NEEDS_OUTREACH.value


def csv_status_to_stage(raw: Optional[str]) -> str:
    """Map a pitch_queue.csv status value to a PMStage string."""
    if not raw:
        return PMStage.NEEDS_OUTREACH.value
    s = raw.strip()
    if s == 'queued':
        return PMStage.QUEUED.value
    if s == 'pitched':
        return PMStage.SENT.value
    if s == 'removed':
        return PMStage.DECLINED.value
    if s == 'not_a_fit':
        # Stage is deprioritized; the not_a_fit flag is the actual exclusion mechanism.
        return PMStage.DEPRIORITIZED.value
    return PMStage.NEEDS_OUTREACH.value


def derive_hs_stage(hs_lead_status: Optional[str], reach_out_1: Optional[date]) -> str:
    """
    Derive PMStage from raw hs_lead_status + reach_out_1. Mirrors get_stage()
    in stages.py so the two never diverge — if stages.py changes, update here too.
    """
    s = hs_lead_status
    if s == 'ATTEMPTED_TO_CONTACT':
        return PMStage.SENT.value
    if s == 'CONNECTED':
        return PMStage.IN_NEGOTIATION.value
    if s == 'BAD_TIMING':
        return PMStage.DEPRIORITIZED.value
    if s == 'UNQUALIFIED':
        return PMStage.DECLINED.value
    if s == 'OPEN_DEAL':
        return PMStage.CONFIRMED.value
    if s in ('OPEN', 'IN_PROGRESS'):
        return PMStage.NEEDS_REVIEW.value
    if s == 'NEW' or reach_out_1 is not None:
        return PMStage.QUEUED.value
    return PMStage.NEEDS_OUTREACH.value


# ── Blank record factory ──────────────────────────────────────────────────────

def _blank(name: str = '') -> dict:
    return {
        'hubspot_id':          None,
        'name':                name,
        'source_hubspot':      False,
        'source_spreadsheet':  False,
        'source_queue_csv':    False,
        'stage':               PMStage.NEEDS_OUTREACH.value,
        'stage_conflict':      False,
        'conflict_note':       None,
        'pitch_type':          None,
        'website':             None,
        'description':         None,
        'reach_out_1':         None,
        'submission_deadline': None,
        'spreadsheet_status':  None,
        'spreadsheet_row':     None,
        'hs_lead_status':      None,
        'queue_csv_status':    None,
        'email_address':       None,
        'not_a_fit':           False,
        'not_a_fit_reason':    None,
    }


# ── In-memory merge state ─────────────────────────────────────────────────────

class _MergeState:
    """
    Collects records from all three source passes into a unified dict before
    any DB writes. Two parallel indices:
      _by_hs_id    — records anchored to a HubSpot ID (primary key)
      _by_name     — records with no known HubSpot ID (name-only dedup)
      _hs_name_idx — normalized name → hubspot_id (fast name-match into HS records)
    """

    def __init__(self):
        self._by_hs_id:    dict[str, dict] = {}
        self._by_name:     dict[str, dict] = {}
        self._hs_name_idx: dict[str, str]  = {}

    def find(self, *, hubspot_id: Optional[str] = None, name: Optional[str] = None) -> Optional[dict]:
        if hubspot_id:
            return self._by_hs_id.get(hubspot_id)
        if name:
            norm  = normalize_name(name)
            hs_id = self._hs_name_idx.get(norm)
            if hs_id:
                return self._by_hs_id.get(hs_id)
            return self._by_name.get(norm)
        return None

    def add_or_get_hs(self, hubspot_id: str, name: str) -> dict:
        record = self._by_hs_id.get(hubspot_id)
        if record is None:
            record = _blank(name)
            record['hubspot_id'] = hubspot_id
            self._by_hs_id[hubspot_id] = record
            self._hs_name_idx[normalize_name(name)] = hubspot_id
        return record

    def add_or_get_name(self, name: str) -> dict:
        """Look up by name. If name matches a HubSpot record (via _hs_name_idx), return that."""
        norm  = normalize_name(name)
        hs_id = self._hs_name_idx.get(norm)
        if hs_id:
            return self._by_hs_id[hs_id]
        record = self._by_name.get(norm)
        if record is None:
            record = _blank(name)
            self._by_name[norm] = record
        return record

    def all_records(self) -> list[dict]:
        return list(self._by_hs_id.values()) + list(self._by_name.values())


# ── Source-specific apply functions ──────────────────────────────────────────

def _apply_hubspot(record: dict, company: HubSpotCompany) -> None:
    record['source_hubspot'] = True
    record['hs_lead_status'] = company.hs_lead_status
    # HubSpot wins on these fields; only fill blanks for the others
    record['website']        = record['website']     or company.website
    record['description']    = record['description'] or company.description
    record['reach_out_1']    = record['reach_out_1'] or company.reach_out_1
    if not record['pitch_type']:
        record['pitch_type'] = 'Festival'


def _apply_spreadsheet(record: dict, row: dict) -> None:
    record['source_spreadsheet'] = True
    record['spreadsheet_status'] = row.get('status')
    record['spreadsheet_row']    = row.get('row')
    record['website']            = record['website']    or row.get('website')
    record['pitch_type']         = record['pitch_type'] or row.get('pitch_type')
    raw_dl = row.get('submission_deadline')
    if raw_dl and not record['submission_deadline']:
        record['submission_deadline'] = _parse_date_value(raw_dl)


def _apply_csv_item(record: dict, item) -> None:
    record['source_queue_csv'] = True
    record['queue_csv_status'] = item.status
    record['email_address']    = record['email_address'] or item.email_address or ''
    record['pitch_type']       = record['pitch_type']    or item.pitch_type
    if item.status == 'not_a_fit':
        record['not_a_fit'] = True
        if item.not_a_fit_reason:
            record['not_a_fit_reason'] = item.not_a_fit_reason
    if item.deadline and not record['reach_out_1']:
        record['reach_out_1'] = item.deadline


# ── Stage computation (Step 4) ────────────────────────────────────────────────

def compute_stage(record: dict) -> None:
    """
    Compute canonical stage and conflict for a fully-merged record dict.
    HubSpot always wins. Called after all three source passes complete.
    Modifies record in place.
    """
    has_hs  = record.get('source_hubspot')
    has_ss  = record.get('source_spreadsheet')
    has_csv = record.get('source_queue_csv')

    if has_hs:
        canonical = derive_hs_stage(
            record.get('hs_lead_status'),
            record.get('reach_out_1'),
        )
        record['stage'] = canonical
        conflict_parts: list[str] = []

        if has_ss and record.get('spreadsheet_status'):
            ss_stage = spreadsheet_status_to_stage(record['spreadsheet_status'])
            if ss_stage != canonical:
                conflict_parts.append(
                    f"Spreadsheet: {record['spreadsheet_status']!r}; "
                    f"HubSpot: {record.get('hs_lead_status')!r}"
                )

        if has_csv and record.get('queue_csv_status'):
            csv_stage = csv_status_to_stage(record['queue_csv_status'])
            if csv_stage != canonical:
                conflict_parts.append(
                    f"Queue CSV: {record['queue_csv_status']!r}; "
                    f"HubSpot: {record.get('hs_lead_status')!r}"
                )

        if conflict_parts:
            record['stage_conflict'] = True
            record['conflict_note']  = '; '.join(conflict_parts)[:500]
        else:
            record['stage_conflict'] = False
            record['conflict_note']  = None
        return

    if has_ss and record.get('spreadsheet_status'):
        ss_stage = spreadsheet_status_to_stage(record['spreadsheet_status'])
        record['stage'] = ss_stage

        if has_csv and record.get('queue_csv_status'):
            csv_stage = csv_status_to_stage(record['queue_csv_status'])
            if csv_stage != ss_stage:
                record['stage_conflict'] = True
                record['conflict_note']  = (
                    f"Spreadsheet: {record['spreadsheet_status']!r}; "
                    f"Queue CSV: {record['queue_csv_status']!r}"
                )[:500]
            else:
                record['stage_conflict'] = False
                record['conflict_note']  = None
        else:
            record['stage_conflict'] = False
            record['conflict_note']  = None
        return

    if has_csv and record.get('queue_csv_status'):
        record['stage']          = csv_status_to_stage(record['queue_csv_status'])
        record['stage_conflict'] = False
        record['conflict_note']  = None
        return

    record['stage']          = PMStage.NEEDS_OUTREACH.value
    record['stage_conflict'] = False
    record['conflict_note']  = None


# ── Spreadsheet XLSX parsing ──────────────────────────────────────────────────

def _detect_columns(header_row: list) -> dict[str, int]:
    """Return {field_name: column_index} using case-insensitive alias matching."""
    lowered = {str(v or '').strip().lower(): i for i, v in enumerate(header_row)}
    result: dict[str, int] = {}
    for field_name, aliases in _COL_ALIASES.items():
        for alias in aliases:
            if alias.lower() in lowered:
                result[field_name] = lowered[alias.lower()]
                break
    return result


def parse_spreadsheet_bytes(raw: bytes) -> list[dict]:
    """
    Parse raw XLSX bytes from the Festival Outreach spreadsheet.
    Returns a list of row dicts with keys from _COL_ALIASES.
    Rows with no name are skipped. Missing columns produce None values.
    Falls back to the active sheet if SPREADSHEET_SHEET_NAME is not found.
    """
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)

    if SPREADSHEET_SHEET_NAME in wb.sheetnames:
        sheet = wb[SPREADSHEET_SHEET_NAME]
    else:
        sheet = wb.active

    all_rows = list(sheet.iter_rows(values_only=True))
    if not all_rows:
        return []

    col_idx = _detect_columns(list(all_rows[0]))
    results: list[dict] = []

    for row_num, row in enumerate(all_rows[1:], start=2):
        def _get(field_name: str) -> Optional[str]:
            idx = col_idx.get(field_name)
            if idx is None or idx >= len(row):
                return None
            val = row[idx]
            if val is None:
                return None
            return str(val).strip() or None

        name = _get('name')
        if not name:
            continue

        results.append({
            'row':                 row_num,
            'name':                name,
            'status':              _get('status'),
            'hubspot_linked':      _get('hubspot_linked'),
            'website':             _get('website'),
            'pitch_type':          _get('pitch_type'),
            'submission_deadline': _get('submission_deadline'),
        })

    return results


def _parse_date_value(value) -> Optional[date]:
    """Parse a date that may arrive as a Python date, datetime, or string from XLSX."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str) and value.strip():
        from app.integrations.pitch_queue import _parse_date
        return _parse_date(value.strip())
    return None


# ── Dropbox download helpers ──────────────────────────────────────────────────

def _try_download_spreadsheet(result: SyncResult) -> Optional[bytes]:
    """
    Download XLSX from Dropbox App folder. Returns None (with warning) if the
    path is not configured or the download fails — the user will move/mirror the
    file into the App folder separately. Never blocks the overall sync.
    """
    from flask import current_app
    from app.integrations.dropbox_sync import DropboxError, _get_access_token, _download_file

    path = current_app.config.get('SPREADSHEET_DROPBOX_PATH', '').strip()
    if not path:
        result.warnings.append(
            'SPREADSHEET_DROPBOX_PATH not configured — spreadsheet source skipped. '
            'Set it to the Dropbox App-folder path once the file is in place.'
        )
        return None
    try:
        token = _get_access_token()
        return _download_file(path, token)
    except DropboxError as e:
        result.warnings.append(f'Spreadsheet download failed ({path!r}): {e} — skipped.')
        return None
    except Exception as e:
        result.warnings.append(f'Spreadsheet download error ({path!r}): {e} — skipped.')
        return None


def _try_download_queue_csv(result: SyncResult) -> Optional[str]:
    """Download pitch_queue.csv. Returns None (with warning) on failure."""
    try:
        from app.integrations.dropbox_sync import get_or_create_queue_csv
        return get_or_create_queue_csv()
    except Exception as e:
        result.warnings.append(f'Queue CSV download failed: {e} — skipped.')
        return None


# ── Public sync API ───────────────────────────────────────────────────────────

def sync_pitch_targets(
    xlsx_bytes:  Optional[bytes] = None,
    csv_content: Optional[str]   = None,
) -> SyncResult:
    """
    Rebuild pitch_targets from all three sources. Atomic: either the table is
    fully replaced with fresh data or an exception is raised and the session
    rolled back (leaving the table unchanged).

    Pass xlsx_bytes/csv_content explicitly in tests to bypass Dropbox calls.
    """
    result = SyncResult()
    state  = _MergeState()

    # ── Step 1: HubSpot companies ──────────────────────────────────────────────
    for company in HubSpotCompany.query.all():
        if company.is_duplicate:
            continue
        record = state.add_or_get_hs(company.hubspot_id, company.name)
        _apply_hubspot(record, company)
        result.hubspot_count += 1

    # ── Step 2: Festival Outreach spreadsheet ──────────────────────────────────
    if xlsx_bytes is None:
        xlsx_bytes = _try_download_spreadsheet(result)

    if xlsx_bytes is not None:
        try:
            ss_rows = parse_spreadsheet_bytes(xlsx_bytes)
        except Exception as e:
            result.warnings.append(f'Spreadsheet parse error: {e} — skipped.')
            ss_rows = []
            result.spreadsheet_skipped = True

        for row in ss_rows:
            name = (row.get('name') or '').strip()
            if not name:
                continue

            hs_linked_raw = row.get('hubspot_linked') or ''
            m = _HS_LINKED_RE.search(hs_linked_raw)
            if m:
                record = state.add_or_get_hs(m.group(1), name)
            else:
                record = state.add_or_get_name(name)

            _apply_spreadsheet(record, row)
            result.spreadsheet_count += 1
    else:
        result.spreadsheet_skipped = True

    # ── Step 3: Queue CSV ──────────────────────────────────────────────────────
    if csv_content is None:
        csv_content = _try_download_queue_csv(result)

    if csv_content:
        from app.integrations.pitch_queue import parse_queue
        for item in parse_queue(csv_content):
            if not item.name:
                continue
            if item.hubspot_id:
                record = state.add_or_get_hs(item.hubspot_id, item.name)
            else:
                record = state.add_or_get_name(item.name)
            _apply_csv_item(record, item)
            result.csv_count += 1

    # ── Step 4: Compute canonical stage + conflicts ────────────────────────────
    all_records = state.all_records()
    for record in all_records:
        compute_stage(record)
        if record['stage_conflict']:
            result.conflict_count += 1
        if record['not_a_fit']:
            result.not_a_fit_count += 1

    # ── Atomic replace ─────────────────────────────────────────────────────────
    now = datetime.utcnow()
    try:
        PitchTarget.query.delete(synchronize_session=False)
        for record in all_records:
            db.session.add(PitchTarget(last_synced_at=now, **record))
        db.session.commit()
    except Exception:
        db.session.rollback()
        raise

    result.total = len(all_records)
    return result
